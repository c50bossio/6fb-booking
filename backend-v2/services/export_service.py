from datetime import datetime, date, timedelta
from typing import List, Optional, Dict, Any
from sqlalchemy.orm import Session
import models
import pandas as pd
import json
import csv
import io
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExportService:
    """Service for handling data exports in various formats"""
    
    def __init__(self):
        self.supported_formats = ['csv', 'excel', 'json', 'pdf']
        self.max_export_records = 50000  # Prevent memory issues
        
    async def export_clients(
        self,
        db: Session,
        format: str = 'csv',
        filters: Optional[Dict[str, Any]] = None,
        include_pii: bool = False,
        user_id: int = None
    ) -> Dict[str, Any]:
        """
        Export client data with filtering and format options
        
        Args:
            db: Database session
            format: Export format (csv, excel, json, pdf)
            filters: Filter criteria
            include_pii: Whether to include personally identifiable information
            user_id: ID of user requesting export (for audit)
            
        Returns:
            Dict with export data and metadata
        """
        logger.info(f"Starting client export: format={format}, filters={filters}, include_pii={include_pii}")
        
        # Build query with filters
        query = db.query(models.Client)
        
        if filters:
            if 'date_from' in filters and filters['date_from']:
                query = query.filter(models.Client.created_at >= filters['date_from'])
            if 'date_to' in filters and filters['date_to']:
                query = query.filter(models.Client.created_at <= filters['date_to'])
            if 'customer_type' in filters and filters['customer_type']:
                query = query.filter(models.Client.customer_type == filters['customer_type'])
            if 'preferred_barber_id' in filters and filters['preferred_barber_id']:
                query = query.filter(models.Client.preferred_barber_id == filters['preferred_barber_id'])
            if 'tags' in filters and filters['tags']:
                query = query.filter(models.Client.tags.contains(filters['tags']))
            if 'min_visits' in filters and filters['min_visits']:
                query = query.filter(models.Client.total_visits >= filters['min_visits'])
            if 'min_spent' in filters and filters['min_spent']:
                query = query.filter(models.Client.total_spent >= filters['min_spent'])
        
        # Limit records to prevent memory issues
        total_count = query.count()
        if total_count > self.max_export_records:
            raise ValueError(f"Export size ({total_count}) exceeds maximum allowed ({self.max_export_records})")
        
        clients = query.order_by(models.Client.created_at.desc()).all()
        
        # Prepare data for export
        client_data = []
        for client in clients:
            client_row = {
                'id': client.id,
                'first_name': client.first_name if include_pii else 'REDACTED',
                'last_name': client.last_name if include_pii else 'REDACTED',
                'email': str(client.email) if include_pii else 'REDACTED',
                'phone': str(client.phone) if (include_pii and client.phone) else 'REDACTED',
                'customer_type': client.customer_type,
                'total_visits': client.total_visits,
                'total_spent': client.total_spent,
                'average_ticket': client.average_ticket,
                'visit_frequency_days': client.visit_frequency_days,
                'no_show_count': client.no_show_count,
                'cancellation_count': client.cancellation_count,
                'referral_count': client.referral_count,
                'first_visit_date': client.first_visit_date.isoformat() if client.first_visit_date else None,
                'last_visit_date': client.last_visit_date.isoformat() if client.last_visit_date else None,
                'tags': client.tags,
                'preferred_barber_id': client.preferred_barber_id,
                'created_at': client.created_at.isoformat(),
                'updated_at': client.updated_at.isoformat() if client.updated_at else None,
            }
            
            if not include_pii:
                client_row['notes'] = 'REDACTED' if client.notes else None
            else:
                client_row['notes'] = str(client.notes) if client.notes else None
                
            client_data.append(client_row)
        
        # Generate export based on format
        if format.lower() == 'csv':
            return await self._export_to_csv(client_data, 'clients')
        elif format.lower() == 'excel':
            return await self._export_clients_to_excel(client_data, filters)
        elif format.lower() == 'json':
            return await self._export_to_json(client_data, 'clients')
        elif format.lower() == 'pdf':
            return await self._export_clients_to_pdf(client_data, filters, include_pii)
        else:
            raise ValueError(f"Unsupported export format: {format}")
    
    async def export_appointments(
        self,
        db: Session,
        format: str = 'csv',
        filters: Optional[Dict[str, Any]] = None,
        include_details: bool = True,
        user_id: int = None
    ) -> Dict[str, Any]:
        """
        Export appointment data with filtering options
        
        Args:
            db: Database session
            format: Export format
            filters: Filter criteria
            include_details: Whether to include detailed information
            user_id: ID of user requesting export
            
        Returns:
            Dict with export data and metadata
        """
        logger.info(f"Starting appointment export: format={format}, filters={filters}")
        
        # Build query with joins for related data
        query = db.query(
            models.Appointment,
            models.User.name.label('user_name'),
            models.User.email.label('user_email'),
            models.Client.first_name.label('client_first_name'),
            models.Client.last_name.label('client_last_name'),
            models.User.name.label('barber_name')  # This will be for barber
        ).outerjoin(
            models.User, models.Appointment.user_id == models.User.id
        ).outerjoin(
            models.Client, models.Appointment.client_id == models.Client.id
        ).outerjoin(
            models.User.alias('barber'), models.Appointment.barber_id == models.User.alias('barber').id
        )
        
        # Apply filters
        if filters:
            if 'date_from' in filters and filters['date_from']:
                query = query.filter(models.Appointment.start_time >= filters['date_from'])
            if 'date_to' in filters and filters['date_to']:
                query = query.filter(models.Appointment.start_time <= filters['date_to'])
            if 'status' in filters and filters['status']:
                if isinstance(filters['status'], list):
                    query = query.filter(models.Appointment.status.in_(filters['status']))
                else:
                    query = query.filter(models.Appointment.status == filters['status'])
            if 'barber_id' in filters and filters['barber_id']:
                query = query.filter(models.Appointment.barber_id == filters['barber_id'])
            if 'service_name' in filters and filters['service_name']:
                query = query.filter(models.Appointment.service_name == filters['service_name'])
            if 'min_price' in filters and filters['min_price']:
                query = query.filter(models.Appointment.price >= filters['min_price'])
            if 'max_price' in filters and filters['max_price']:
                query = query.filter(models.Appointment.price <= filters['max_price'])
        
        # Limit records
        total_count = query.count()
        if total_count > self.max_export_records:
            raise ValueError(f"Export size ({total_count}) exceeds maximum allowed ({self.max_export_records})")
        
        appointments = query.order_by(models.Appointment.start_time.desc()).all()
        
        # Prepare data
        appointment_data = []
        for row in appointments:
            appointment = row[0]  # The Appointment object
            
            appointment_row = {
                'id': appointment.id,
                'user_name': row.user_name,
                'user_email': row.user_email,
                'client_name': f"{row.client_first_name or ''} {row.client_last_name or ''}".strip() or 'N/A',
                'barber_name': row.barber_name or 'Unassigned',
                'service_name': appointment.service_name,
                'start_time': appointment.start_time.isoformat(),
                'duration_minutes': appointment.duration_minutes,
                'price': appointment.price,
                'status': appointment.status,
                'created_at': appointment.created_at.isoformat(),
            }
            
            if include_details:
                appointment_row.update({
                    'notes': appointment.notes or '',
                    'buffer_time_before': appointment.buffer_time_before or 0,
                    'buffer_time_after': appointment.buffer_time_after or 0,
                    'google_event_id': appointment.google_event_id or '',
                })
            
            appointment_data.append(appointment_row)
        
        # Generate export
        if format.lower() == 'csv':
            return await self._export_to_csv(appointment_data, 'appointments')
        elif format.lower() == 'excel':
            return await self._export_appointments_to_excel(appointment_data, filters)
        elif format.lower() == 'json':
            return await self._export_to_json(appointment_data, 'appointments')
        elif format.lower() == 'pdf':
            return await self._export_appointments_to_pdf(appointment_data, filters)
        else:
            raise ValueError(f"Unsupported export format: {format}")
    
    async def export_analytics(
        self,
        db: Session,
        format: str = 'excel',
        date_range: Optional[Dict[str, date]] = None,
        include_charts: bool = True,
        user_id: int = None
    ) -> Dict[str, Any]:
        """
        Export analytics data with charts and summaries
        
        Args:
            db: Database session
            format: Export format (excel recommended for charts)
            date_range: Date range for analytics
            include_charts: Whether to include charts in export
            user_id: ID of user requesting export
            
        Returns:
            Dict with export data and metadata
        """
        logger.info(f"Starting analytics export: format={format}, date_range={date_range}")
        
        # Set default date range if not provided
        if not date_range:
            end_date = date.today()
            start_date = end_date - timedelta(days=90)  # Last 3 months
            date_range = {'start_date': start_date, 'end_date': end_date}
        
        # Get analytics data from existing service
        from services import analytics_service
        
        # Revenue analytics
        revenue_data = analytics_service.get_revenue_analytics(
            db, 
            start_date=date_range['start_date'], 
            end_date=date_range['end_date']
        )
        
        # Performance analytics
        performance_data = analytics_service.get_performance_analytics(
            db,
            start_date=date_range['start_date'],
            end_date=date_range['end_date']
        )
        
        # Client analytics
        client_analytics = analytics_service.get_client_analytics(
            db,
            start_date=date_range['start_date'],
            end_date=date_range['end_date']
        )
        
        # Six Figure Barber metrics
        try:
            six_figure_metrics = analytics_service.get_six_figure_barber_metrics(db)
        except Exception as e:
            logger.warning(f"Could not get Six Figure Barber metrics: {e}")
            six_figure_metrics = None
        
        analytics_data = {
            'revenue': revenue_data,
            'performance': performance_data,
            'clients': client_analytics,
            'six_figure_barber': six_figure_metrics,
            'date_range': date_range
        }
        
        # Generate export
        if format.lower() == 'excel':
            return await self._export_analytics_to_excel(analytics_data, include_charts)
        elif format.lower() == 'json':
            return await self._export_to_json(analytics_data, 'analytics')
        elif format.lower() == 'pdf':
            return await self._export_analytics_to_pdf(analytics_data)
        else:
            raise ValueError(f"Analytics export not supported for format: {format}")
    
    async def custom_export(
        self,
        db: Session,
        query_config: Dict[str, Any],
        format: str = 'csv',
        user_id: int = None
    ) -> Dict[str, Any]:
        """
        Custom export based on user-defined configuration
        
        Args:
            db: Database session
            query_config: Configuration defining what to export
            format: Export format
            user_id: ID of user requesting export
            
        Returns:
            Dict with export data and metadata
        """
        logger.info(f"Starting custom export: config={query_config}, format={format}")
        
        # Parse query configuration
        table = query_config.get('table')
        fields = query_config.get('fields', [])
        filters = query_config.get('filters', {})
        joins = query_config.get('joins', [])
        order_by = query_config.get('order_by', 'created_at')
        limit = min(query_config.get('limit', 1000), self.max_export_records)
        
        if not table or not fields:
            raise ValueError("Custom export requires 'table' and 'fields' in query_config")
        
        # Build dynamic query (simplified version - in production, use a query builder)
        if table == 'clients':
            data = await self.export_clients(db, format, filters, user_id=user_id)
        elif table == 'appointments':
            data = await self.export_appointments(db, format, filters, user_id=user_id)
        else:
            raise ValueError(f"Custom export not supported for table: {table}")
        
        return data
    
    # Format-specific export methods
    
    async def _export_to_csv(self, data: List[Dict], data_type: str) -> Dict[str, Any]:
        """Export data to CSV format"""
        if not data:
            return {'content': '', 'filename': f'{data_type}_export.csv', 'mime_type': 'text/csv'}
        
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
        
        content = output.getvalue()
        output.close()
        
        return {
            'content': base64.b64encode(content.encode()).decode(),
            'filename': f'{data_type}_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv',
            'mime_type': 'text/csv',
            'encoding': 'base64'
        }
    
    async def _export_to_json(self, data: Any, data_type: str) -> Dict[str, Any]:
        """Export data to JSON format"""
        json_content = json.dumps(data, indent=2, default=str)
        
        return {
            'content': base64.b64encode(json_content.encode()).decode(),
            'filename': f'{data_type}_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json',
            'mime_type': 'application/json',
            'encoding': 'base64'
        }
    
    async def _export_clients_to_excel(self, data: List[Dict], filters: Dict = None) -> Dict[str, Any]:
        """Export clients to Excel with formatting and summary"""
        wb = Workbook()
        
        # Main data sheet
        ws = wb.active
        ws.title = "Clients"
        
        if data:
            # Convert to DataFrame for easier manipulation
            df = pd.DataFrame(data)
            
            # Add headers with formatting
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Format headers
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add summary sheet
            summary_ws = wb.create_sheet("Summary")
            summary_data = self._generate_client_summary(df)
            
            summary_ws.append(["Client Export Summary"])
            summary_ws.append([])
            for key, value in summary_data.items():
                summary_ws.append([key.replace('_', ' ').title(), value])
            
            # Format summary sheet
            summary_ws['A1'].font = Font(bold=True, size=14)
            for row in summary_ws.iter_rows(min_row=3):
                row[0].font = Font(bold=True)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return {
            'content': base64.b64encode(output.getvalue()).decode(),
            'filename': f'clients_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            'mime_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'encoding': 'base64'
        }
    
    async def _export_appointments_to_excel(self, data: List[Dict], filters: Dict = None) -> Dict[str, Any]:
        """Export appointments to Excel with formatting and charts"""
        wb = Workbook()
        
        # Main data sheet
        ws = wb.active
        ws.title = "Appointments"
        
        if data:
            df = pd.DataFrame(data)
            
            # Add data to worksheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Format headers
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add summary sheet with charts
            summary_ws = wb.create_sheet("Summary")
            self._add_appointment_summary_charts(summary_ws, df)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return {
            'content': base64.b64encode(output.getvalue()).decode(),
            'filename': f'appointments_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            'mime_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'encoding': 'base64'
        }
    
    async def _export_analytics_to_excel(self, data: Dict, include_charts: bool = True) -> Dict[str, Any]:
        """Export analytics to Excel with charts and summaries"""
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Revenue sheet
        if 'revenue' in data and data['revenue']:
            revenue_ws = wb.create_sheet("Revenue Analytics")
            self._add_revenue_sheet(revenue_ws, data['revenue'], include_charts)
        
        # Performance sheet
        if 'performance' in data and data['performance']:
            perf_ws = wb.create_sheet("Performance Analytics")
            self._add_performance_sheet(perf_ws, data['performance'])
        
        # Client analytics sheet
        if 'clients' in data and data['clients']:
            clients_ws = wb.create_sheet("Client Analytics")
            self._add_client_analytics_sheet(clients_ws, data['clients'])
        
        # Six Figure Barber sheet
        if 'six_figure_barber' in data and data['six_figure_barber']:
            sfb_ws = wb.create_sheet("Six Figure Barber")
            self._add_six_figure_barber_sheet(sfb_ws, data['six_figure_barber'])
        
        # Executive summary sheet
        summary_ws = wb.create_sheet("Executive Summary")
        wb.active = summary_ws  # Make this the first sheet
        self._add_executive_summary(summary_ws, data)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return {
            'content': base64.b64encode(output.getvalue()).decode(),
            'filename': f'analytics_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            'mime_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'encoding': 'base64'
        }
    
    def _generate_client_summary(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Generate summary statistics for client data"""
        return {
            'total_clients': len(df),
            'active_clients': len(df[df['total_visits'] > 0]),
            'average_visits': df['total_visits'].mean(),
            'average_spent': df['total_spent'].mean(),
            'total_revenue': df['total_spent'].sum(),
            'customer_types': df['customer_type'].value_counts().to_dict(),
            'no_show_rate': (df['no_show_count'].sum() / df['total_visits'].sum() * 100) if df['total_visits'].sum() > 0 else 0
        }
    
    def _add_appointment_summary_charts(self, ws, df: pd.DataFrame):
        """Add summary charts to appointment export"""
        # Status distribution
        status_counts = df['status'].value_counts()
        
        ws.append(["Appointment Summary"])
        ws.append([])
        ws.append(["Status", "Count"])
        
        for status, count in status_counts.items():
            ws.append([status, count])
        
        # Add pie chart for status distribution
        pie = PieChart()
        labels = Reference(ws, min_col=1, min_row=4, max_row=3+len(status_counts))
        data = Reference(ws, min_col=2, min_row=3, max_row=3+len(status_counts))
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Appointments by Status"
        
        ws.add_chart(pie, "E3")
    
    def _add_revenue_sheet(self, ws, revenue_data: Dict, include_charts: bool):
        """Add revenue analytics to worksheet"""
        ws.append(["Revenue Analytics"])
        ws.append([])
        
        # Add summary metrics
        if 'total_revenue' in revenue_data:
            ws.append(["Total Revenue", revenue_data['total_revenue']])
        if 'growth_metrics' in revenue_data:
            ws.append(["Revenue Growth %", revenue_data['growth_metrics'].get('revenue_growth', 0)])
            ws.append(["Ticket Growth %", revenue_data['growth_metrics'].get('ticket_growth', 0)])
        
        # Add period data if available
        if 'revenue_by_period' in revenue_data and revenue_data['revenue_by_period']:
            ws.append([])
            ws.append(["Period", "Revenue", "Appointments", "Average Ticket"])
            for period_data in revenue_data['revenue_by_period']:
                ws.append([
                    period_data.get('period', ''),
                    period_data.get('revenue', 0),
                    period_data.get('appointments', 0),
                    period_data.get('average_ticket', 0)
                ])
            
            if include_charts:
                # Add line chart for revenue trend
                chart = LineChart()
                data = Reference(ws, min_col=2, min_row=7, max_row=6+len(revenue_data['revenue_by_period']))
                chart.add_data(data, titles_from_data=True)
                chart.title = "Revenue Trend"
                ws.add_chart(chart, "F3")
    
    def _add_performance_sheet(self, ws, performance_data: Dict):
        """Add performance analytics to worksheet"""
        ws.append(["Performance Analytics"])
        ws.append([])
        
        if 'efficiency_metrics' in performance_data:
            metrics = performance_data['efficiency_metrics']
            ws.append(["Efficiency Metrics"])
            ws.append(["Utilization Rate %", metrics.get('utilization_rate', 0)])
            ws.append(["Avg Appointment Duration", metrics.get('average_appointment_duration', 0)])
            ws.append(["Daily Capacity", metrics.get('daily_capacity', 0)])
            ws.append([])
        
        if 'client_metrics' in performance_data:
            metrics = performance_data['client_metrics']
            ws.append(["Client Metrics"])
            ws.append(["New Clients", metrics.get('new_clients_count', 0)])
            ws.append(["Returning Clients", metrics.get('returning_clients_count', 0)])
            ws.append(["Retention Rate %", metrics.get('client_retention_rate', 0)])
            ws.append([])
        
        if 'financial_metrics' in performance_data:
            metrics = performance_data['financial_metrics']
            ws.append(["Financial Metrics"])
            ws.append(["Revenue per Hour", metrics.get('revenue_per_hour', 0)])
            ws.append(["Revenue per Client", metrics.get('revenue_per_client', 0)])
    
    def _add_client_analytics_sheet(self, ws, client_data: Dict):
        """Add client analytics to worksheet"""
        ws.append(["Client Analytics"])
        ws.append([])
        
        # Add client summary data
        for key, value in client_data.items():
            if isinstance(value, dict):
                ws.append([key.replace('_', ' ').title()])
                for sub_key, sub_value in value.items():
                    ws.append([f"  {sub_key.replace('_', ' ').title()}", sub_value])
                ws.append([])
            else:
                ws.append([key.replace('_', ' ').title(), value])
    
    def _add_six_figure_barber_sheet(self, ws, sfb_data: Dict):
        """Add Six Figure Barber metrics to worksheet"""
        ws.append(["Six Figure Barber Metrics"])
        ws.append([])
        
        if 'current_performance' in sfb_data:
            perf = sfb_data['current_performance']
            ws.append(["Current Performance"])
            ws.append(["Monthly Revenue", perf.get('monthly_revenue', 0)])
            ws.append(["Annual Revenue Projection", perf.get('annual_revenue_projection', 0)])
            ws.append(["Average Ticket", perf.get('average_ticket', 0)])
            ws.append(["Utilization Rate %", perf.get('utilization_rate', 0)])
            ws.append([])
        
        if 'targets' in sfb_data:
            targets = sfb_data['targets']
            ws.append(["Targets"])
            ws.append(["Annual Income Target", targets.get('annual_income_target', 0)])
            ws.append(["Monthly Revenue Target", targets.get('monthly_revenue_target', 0)])
            ws.append(["Revenue Gap", targets.get('revenue_gap', 0)])
            ws.append(["On Track", "Yes" if targets.get('on_track', False) else "No"])
            ws.append([])
        
        if 'action_items' in sfb_data:
            ws.append(["Action Items"])
            for item in sfb_data['action_items']:
                ws.append([f"• {item}"])
    
    def _add_executive_summary(self, ws, data: Dict):
        """Add executive summary to worksheet"""
        ws.append(["6FB Booking Analytics - Executive Summary"])
        ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws.append([])
        
        # Date range
        if 'date_range' in data:
            date_range = data['date_range']
            ws.append([f"Report Period: {date_range['start_date']} to {date_range['end_date']}"])
            ws.append([])
        
        # Key metrics summary
        ws.append(["Key Performance Indicators"])
        
        if 'revenue' in data and data['revenue']:
            revenue = data['revenue']
            if 'total_revenue' in revenue:
                ws.append(["Total Revenue", f"${revenue['total_revenue']:,.2f}"])
        
        if 'performance' in data and data['performance']:
            perf = data['performance']
            if 'efficiency_metrics' in perf:
                eff = perf['efficiency_metrics']
                ws.append(["Utilization Rate", f"{eff.get('utilization_rate', 0):.1f}%"])
            if 'client_metrics' in perf:
                client = perf['client_metrics']
                ws.append(["Client Retention Rate", f"{client.get('client_retention_rate', 0):.1f}%"])
        
        # Format header
        ws['A1'].font = Font(bold=True, size=16)
        ws['A2'].font = Font(italic=True)
        ws['A5'].font = Font(bold=True, size=12)
        
        for row in ws.iter_rows(min_row=6):
            if row[0].value:
                row[0].font = Font(bold=True)

# Global service instance
export_service = ExportService()